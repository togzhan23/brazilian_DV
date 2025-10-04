
import os
import uuid
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine
import psycopg2
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

DB_USER = "postgres"
DB_PASS = "0000"
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "brazilian"

CONN_STR = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

#helper functions 
def get_engine():
    return create_engine(CONN_STR)

def fetch_sql(sql, engine):
    """Fetch SQL into pandas DataFrame"""
    return pd.read_sql_query(sql, engine)

def save_matplotlib(fig, filename):
    fig.tight_layout()
    fig.savefig(filename)
    plt.close(fig)

def console_report(name, df, chart_type, description, path=None):
    print("-" * 60)
    print(f"Chart: {name}")
    print(f"Type: {chart_type}")
    print(f"Rows returned: {len(df)}")
    print(f"Summary: {description}")
    if path:
        print(f"Saved to: {path}")
    print("-" * 60)


# SQL queries joins, multi tables 
# 1. Pie chart: revenue distribution by product category (top categories)
SQL_PIE = """
SELECT pct.product_category_name_english AS category,
       SUM(oi.price + oi.freight_value) AS revenue
FROM order_items oi
JOIN products p ON oi.product_id = p.product_id
JOIN product_category_translation pct ON p.product_category_name = pct.product_category_name
JOIN orders o ON oi.order_id = o.order_id
WHERE o.order_status = 'delivered'
GROUP BY pct.product_category_name_english
ORDER BY revenue DESC;
"""

# 2) Bar chart: top 10 seller cities by revenue
SQL_BAR = """
SELECT s.seller_city AS seller_city,
       SUM(oi.price + oi.freight_value) AS revenue
FROM order_items oi
JOIN sellers s ON oi.seller_id = s.seller_id
JOIN orders o ON oi.order_id = o.order_id
JOIN products p ON oi.product_id = p.product_id
WHERE o.order_status = 'delivered'
GROUP BY s.seller_city
ORDER BY revenue DESC
LIMIT 10;
"""

# 3) Horizontal bar: average order value by customer state (top 10 states)
SQL_HBAR = """
WITH order_totals AS (
  SELECT oi.order_id, SUM(oi.price + oi.freight_value) AS order_total
  FROM order_items oi
  GROUP BY oi.order_id
)
SELECT c.customer_state,
       AVG(ot.order_total) AS avg_order_value
FROM order_totals ot
JOIN orders o ON ot.order_id = o.order_id
JOIN customers c ON o.customer_id = c.customer_id
WHERE o.order_status = 'delivered'
GROUP BY c.customer_state
ORDER BY avg_order_value DESC
LIMIT 10;
"""

# 4) Line chart: monthly revenue over time
SQL_LINE = """
SELECT to_char(date_trunc('month', o.order_purchase_timestamp), 'YYYY-MM') AS month,
       SUM(oi.price + oi.freight_value) AS revenue
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
JOIN products p ON oi.product_id = p.product_id
GROUP BY month
ORDER BY month;
"""

# 5) Histogram: distribution of order totals
SQL_HIST = """
SELECT oi.order_id,
       SUM(oi.price + oi.freight_value) AS order_total
FROM order_items oi
JOIN orders o ON oi.order_id = o.order_id
JOIN customers c ON o.customer_id = c.customer_id
GROUP BY oi.order_id;
"""

# 6) Scatter: delivery days vs review score (already readable)
SQL_SCATTER = """
SELECT orv.review_score,
       EXTRACT(EPOCH FROM (o.order_delivered_customer_date - o.order_purchase_timestamp))/86400.0 AS delivery_days
FROM order_reviews orv
JOIN orders o ON orv.order_id = o.order_id
JOIN customers c ON o.customer_id = c.customer_id
WHERE o.order_delivered_customer_date IS NOT NULL
  AND orv.review_score IS NOT NULL;
"""

# Time slider query (for Plotly interactive) - monthly revenue by state
SQL_TIME_SLIDER = """
SELECT to_char(date_trunc('month', o.order_purchase_timestamp), 'YYYY-MM') AS month,
       c.customer_state,
       SUM(oi.price + oi.freight_value) AS revenue
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
JOIN customers c ON o.customer_id = c.customer_id
GROUP BY month, c.customer_state
ORDER BY month, c.customer_state;
"""

# plotting and export logic 
def build_pie(engine):
    df = fetch_sql(SQL_PIE, engine)
    if df.empty:
        print("Pie query returned 0 rows.")
        return df
    # keep top N categories and group rest into 'Other'
    TOP_N = 8
    df_top = df.head(TOP_N).copy()
    others_sum = df[TOP_N:]["revenue"].sum()
    if others_sum > 0:
        df_top = pd.concat([df_top, pd.DataFrame([{"category":"Other","revenue":others_sum}])], ignore_index=True)
    fig, ax = plt.subplots(figsize=(8,6))
    ax.pie(df_top["revenue"], labels=df_top["category"], autopct="%1.1f%%")
    ax.set_title("Revenue distribution by product category (top categories)")
    out = os.path.join(CHARTS_DIR, "pie_category_revenue.png")
    save_matplotlib(fig, out)
    console_report("Revenue by Category", df, "Pie chart", "Shows share of revenue by product category (delivered orders)", out)
    return df

def build_bar(engine):
    df = fetch_sql(SQL_BAR, engine)
    if df.empty:
        print("Bar query returned 0 rows.")
        return df
    fig, ax = plt.subplots(figsize=(10,6))
    ax.bar(df['seller_city'].astype(str), df['revenue'])
    ax.set_xlabel("Seller City")
    ax.set_ylabel("Revenue (price + freight)")
    ax.set_title("Top 10 seller cities by revenue (delivered orders)")
    plt.xticks(rotation=45, ha='right')
    out = os.path.join(CHARTS_DIR, "bar_top_seller_cities.png")
    save_matplotlib(fig, out)
    console_report("Top Seller Cities", df, "Bar chart", "Top 10 seller cities by revenue", out)
    return df


def build_hbar(engine):
    df = fetch_sql(SQL_HBAR, engine)
    if df.empty:
        print("HBar query returned 0 rows.")
        return df
    # horizontal bar
    fig, ax = plt.subplots(figsize=(8,6))
    ax.barh(df['customer_state'].astype(str), df['avg_order_value'])
    ax.set_xlabel("Average Order Value")
    ax.set_ylabel("Customer State")
    ax.set_title("Average order value by customer state (top 10)")
    out = os.path.join(CHARTS_DIR, "hbar_avg_order_by_state.png")
    save_matplotlib(fig, out)
    console_report("Avg Order by State", df, "Horizontal bar chart", "Average order value per customer state (delivered orders)", out)
    return df

def build_line(engine):
    df = fetch_sql(SQL_LINE, engine)
    if df.empty:
        print("Line query returned 0 rows.")
        return df
    # ensure month is datetime for plotting
    df['month'] = pd.to_datetime(df['month'])
    fig, ax = plt.subplots(figsize=(10,6))
    ax.plot(df['month'], df['revenue'], marker='o')
    ax.set_xlabel("Month")
    ax.set_ylabel("Revenue")
    ax.set_title("Monthly revenue trend (all products)")
    out = os.path.join(CHARTS_DIR, "line_monthly_revenue.png")
    save_matplotlib(fig, out)
    console_report("Monthly Revenue", df, "Line chart", "Shows revenue trend by month", out)
    return df

def build_hist(engine):
    df = fetch_sql(SQL_HIST, engine)
    if df.empty:
        print("Hist query returned 0 rows.")
        return df
    fig, ax = plt.subplots(figsize=(8,6))
    ax.hist(df['order_total'], bins=50)
    ax.set_xlabel("Order total (price + freight)")
    ax.set_ylabel("Number of orders")
    ax.set_title("Distribution of order totals")
    out = os.path.join(CHARTS_DIR, "hist_order_totals.png")
    save_matplotlib(fig, out)
    console_report("Order total distribution", df, "Histogram", "Distribution of order totals across all orders", out)
    return df

def build_scatter(engine):
    df = fetch_sql(SQL_SCATTER, engine)
    df = df.dropna()
    if df.empty:
        print("Scatter query returned 0 rows.")
        return df
    fig, ax = plt.subplots(figsize=(8,6))
    ax.scatter(df['delivery_days'], df['review_score'], alpha=0.6)
    ax.set_xlabel("Delivery days")
    ax.set_ylabel("Review score")
    ax.set_title("Delivery time vs Review score")
    out = os.path.join(CHARTS_DIR, "scatter_delivery_vs_review.png")
    save_matplotlib(fig, out)
    console_report("Delivery vs Review", df, "Scatter plot", "Shows relation between delivery days and customer review score", out)
    return df

#  --------------------------------------------------------------------------

def build_plotly_time_slider(engine):
    df = fetch_sql(SQL_TIME_SLIDER, engine)
    if df.empty:
        print("Plotly time-slider query returned 0 rows.")
        return df
    # month column already 'YYYY-MM' string
    # We'll show top states per month as bar animation (or scatter)
    # Use plotly.express bar with animation_frame
    fig = px.bar(df, x='customer_state', y='revenue', animation_frame='month', animation_group='customer_state',
                 labels={'customer_state':'State','revenue':'Revenue'}, title="Monthly revenue by state (use slider)")
    fig.update_layout(xaxis={'categoryorder':'total descending'})
    import plotly.io as pio
    pio.renderers.default = "browser"
    fig.show()
    print("interactive Plotly chart with time slider.")
    return df
#  --------------------------------------------------------------------------

#  Excel export 
def export_to_excel(dataframes_dict, filename):
    """
    dataframes_dict: dict of sheet_name -> DataFrame
    filename: path to save file (xlsx)
    Applies:
      - freeze top row
      - auto filter
      - colour scale for numeric columns
    """
    out_path = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
    # formatting
    wb = load_workbook(out_path)
    total_rows = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # freeze header row
        ws.freeze_panes = "A2"
        # autofiliter (use dimensions after write)
        ws.auto_filter.ref = ws.dimensions
        # find numeric columns by header detection in pandas df may be easier:
        # apply a color scale to whole numeric columns (starting from row 2)
        # We approximate: for each column letter, check first data cell if it is numeric
        # Determine col letters and ranges
        max_col = ws.max_column
        max_row = ws.max_row
        total_rows += (max_row - 1)
        # apply color scale to each column that looks numeric
        for col in ws.iter_cols(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            col_letter = col[0].column_letter
            # check if the first value is numeric
            sample_val = col[0].value
            if isinstance(sample_val, (int, float)):
                rng = f"{col_letter}2:{col_letter}{max_row}"
                rule = ColorScaleRule(start_type="min", start_color="FFAA0000",
                                      mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                                      end_type="max", end_color="FF00AA00")
                ws.conditional_formatting.add(rng, rule)
    wb.save(out_path)
    print(f"Created file {out_path}, sheets: {len(wb.sheetnames)}, total data rows (approx): {total_rows}")



#  --------------------------------------------------------------------------



# main execution 
def main():
    engine = get_engine()

    # Build and save 6 charts
    df_pie = build_pie(engine)
    df_bar = build_bar(engine)
    df_hbar = build_hbar(engine)
    df_line = build_line(engine)
    df_hist = build_hist(engine)
    df_scatter = build_scatter(engine)

    # Interactive plotly slider 
    df_time = build_plotly_time_slider(engine)

    # Export some results to Excel with formatting
    to_export = {
        "top_categories": df_pie.head(50) if df_pie is not None else pd.DataFrame(),
        "top_sellers": df_bar if df_bar is not None else pd.DataFrame(),
        "monthly_revenue": df_line if df_line is not None else pd.DataFrame()
    }
    export_to_excel(to_export, "olist_report.xlsx")

    print("\nAll done. Charts saved to ./charts/, exports to ./exports/")

    # OPTIONAL: demonstrate inserting a row and regenerating one chart:
    # demo_id = insert_demo_order()
    # if demo_id:
    #     print("Re-run the relevant chart query (e.g., build_line) to see changes after insert.")
    #     build_line(engine)

if __name__ == "__main__":
    main()
